#!/usr/bin/env python3
"""
Script to analyze VSME Excel Template structure
"""
import sys
try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

def analyze_excel(file_path):
    """Analyze the Excel file structure"""
    print(f"Analyzing: {file_path}\n")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Get all sheet names
        print(f"\n📊 Found {len(wb.sheetnames)} sheets:")
        print("-" * 80)
        
        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"\n{idx}. Sheet: '{sheet_name}'")
            ws = wb[sheet_name]
            
            # Get sheet dimensions
            if ws.max_row > 0 and ws.max_column > 0:
                print(f"   Dimensions: {ws.max_row} rows x {ws.max_column} columns")
                
                # Find header rows (usually rows with text in bold or specific patterns)
                print(f"   First 10 rows analysis:")
                for row_idx in range(1, min(11, ws.max_row + 1)):
                    row_data = []
                    for col_idx in range(1, min(6, ws.max_column + 1)):  # First 5 columns
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        if value:
                            # Truncate long values
                            cell_value = str(value)[:50] if value else ""
                            row_data.append(cell_value)
                    
                    if row_data:
                        print(f"      Row {row_idx}: {', '.join(row_data)}")
                
                # Look for data fields (cells with labels)
                print(f"\n   Field Analysis (looking for labels):")
                fields = []
                for row_idx in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
                    for col_idx in range(1, min(ws.max_column + 1, 10)):  # First 10 columns
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value and isinstance(cell.value, str):
                            value = str(cell.value).strip()
                            # Look for field-like patterns
                            if len(value) > 3 and len(value) < 100:
                                # Check if it looks like a field label
                                if any(keyword in value.lower() for keyword in 
                                       ['name', 'address', 'date', 'number', 'code', 'id', 
                                        'company', 'country', 'city', 'email', 'phone',
                                        'module', 'section', 'disclosure', 'report']):
                                    fields.append({
                                        'row': row_idx,
                                        'col': col_idx,
                                        'value': value,
                                        'sheet': sheet_name
                                    })
                
                if fields:
                    print(f"   Found {len(fields)} potential fields (showing first 20):")
                    for field in fields[:20]:
                        print(f"      Row {field['row']}, Col {field['col']}: {field['value']}")
                else:
                    print("   No obvious field labels found in first 100 rows")
            else:
                print("   Empty sheet")
            
            print()
        
        print("=" * 80)
        print("\n📋 Summary:")
        print(f"Total sheets: {len(wb.sheetnames)}")
        print(f"Sheet names: {', '.join(wb.sheetnames)}")
        
        # Look for "Basic" related sheets
        basic_sheets = [s for s in wb.sheetnames if 'basic' in s.lower()]
        if basic_sheets:
            print(f"\n🎯 Basic Report related sheets: {', '.join(basic_sheets)}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    excel_file = "VSME-Digital-Template-1.1.0.xlsx"
    analyze_excel(excel_file)


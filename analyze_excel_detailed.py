#!/usr/bin/env python3
"""
Detailed analysis of VSME Excel Template - Basic Report fields
"""
import sys
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.utils import get_column_letter

def analyze_sheet_detailed(ws, sheet_name):
    """Analyze a sheet in detail for form fields"""
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*80}")
    
    fields = []
    
    # Look through rows for field definitions
    # Typically fields are in column C (3) with labels, and data goes in columns D, E, F, etc.
    for row_idx in range(1, min(ws.max_row + 1, 300)):  # Check first 300 rows
        row_data = {}
        
        # Check column C (3) for field labels
        label_cell = ws.cell(row=row_idx, column=3)
        if label_cell.value and isinstance(label_cell.value, str):
            label = str(label_cell.value).strip()
            
            # Skip empty or very short labels
            if len(label) < 5:
                continue
            
            # Look for field patterns (questions, disclosures, etc.)
            if any(keyword in label.lower() for keyword in [
                'name', 'identifier', 'currency', 'date', 'year', 'month', 'day',
                'country', 'city', 'address', 'email', 'phone', 'number',
                'employee', 'energy', 'consumption', 'emission', 'ghg', 'co2',
                'contract', 'gender', 'accident', 'safety', 'turnover',
                'conviction', 'fine', 'corruption', 'revenue', 'disclosure',
                'report', 'period', 'starting', 'ending', 'nace', 'sector',
                'subsidiary', 'basis', 'preparation', 'module'
            ]) or label.startswith('B') or '[Always to be reported]' in label or '[If applicable]' in label:
                
                # Get the data type hint from nearby cells
                data_type = 'text'
                required = '[Always to be reported]' in label or '[Alw' in label
                optional = '[If applicable]' in label
                
                # Check column D, E, F for example values or data types
                for col_idx in range(4, min(8, ws.max_column + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        value_str = str(cell.value).strip()
                        if value_str and value_str not in ['-', 'N/A', 'TBD']:
                            # Try to infer data type
                            if any(x in value_str.lower() for x in ['yes', 'no', 'true', 'false']):
                                data_type = 'boolean'
                            elif any(x in value_str for x in ['http', 'www.', '.com']):
                                data_type = 'url'
                            elif '@' in value_str:
                                data_type = 'email'
                            elif value_str.replace('.', '').replace('-', '').isdigit():
                                data_type = 'number'
                            break
                
                # Check for dropdown/select options (usually in rows below)
                options = []
                if row_idx + 1 <= ws.max_row:
                    # Look for options in the next few rows
                    for check_row in range(row_idx + 1, min(row_idx + 10, ws.max_row + 1)):
                        option_cell = ws.cell(row=check_row, column=4)
                        if option_cell.value and isinstance(option_cell.value, str):
                            option = str(option_cell.value).strip()
                            if option and len(option) < 100:
                                # Check if it's a valid option (not a header or instruction)
                                if not any(skip in option.lower() for skip in [
                                    'current', 'previous', 'reporting', 'period', 'year', 'date',
                                    'please', 'select', 'enter', 'provide', 'indicates'
                                ]):
                                    options.append(option)
                
                field = {
                    'row': row_idx,
                    'label': label,
                    'data_type': data_type,
                    'required': required,
                    'optional': optional,
                    'options': options[:10] if options else None,  # Limit to 10 options
                    'sheet': sheet_name
                }
                
                fields.append(field)
    
    return fields

def analyze_basic_report_sheets(wb):
    """Analyze sheets relevant to Basic Report"""
    
    # Basic Report sheets based on VSME standard
    basic_sheets = [
        'General Information',
        'Environmental Disclosures',
        'Social Disclosures',
        'Governance Disclosures'
    ]
    
    all_fields = {}
    
    for sheet_name in basic_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            fields = analyze_sheet_detailed(ws, sheet_name)
            all_fields[sheet_name] = fields
            
            print(f"\nFound {len(fields)} fields in {sheet_name}")
            print("-" * 80)
            
            # Group fields by category
            for field in fields[:30]:  # Show first 30 fields
                req_str = "[REQUIRED]" if field['required'] else "[OPTIONAL]" if field['optional'] else ""
                type_str = f"Type: {field['data_type']}"
                options_str = f"Options: {', '.join(field['options'])}" if field['options'] else ""
                
                print(f"  Row {field['row']:3d}: {req_str} {field['label'][:60]}")
                if options_str:
                    print(f"           {options_str[:70]}")
        else:
            print(f"\n⚠️  Sheet '{sheet_name}' not found in workbook")
    
    return all_fields

def main():
    excel_file = "VSME-Digital-Template-1.1.0.xlsx"
    
    print("=" * 80)
    print("VSME Excel Template - Basic Report Field Analysis")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        print(f"\n📊 Workbook has {len(wb.sheetnames)} sheets")
        print(f"Sheet names: {', '.join(wb.sheetnames)}")
        
        # Analyze Basic Report sheets
        fields = analyze_basic_report_sheets(wb)
        
        # Summary
        print("\n" + "=" * 80)
        print("SUMMARY")
        print("=" * 80)
        total_fields = 0
        required_fields = 0
        for sheet_name, sheet_fields in fields.items():
            total_fields += len(sheet_fields)
            required_fields += sum(1 for f in sheet_fields if f['required'])
            print(f"{sheet_name}: {len(sheet_fields)} fields ({sum(1 for f in sheet_fields if f['required'])} required)")
        
        print(f"\nTotal Basic Report fields: {total_fields}")
        print(f"Total required fields: {required_fields}")
        
        # Generate field mapping structure
        print("\n" + "=" * 80)
        print("FIELD MAPPING STRUCTURE (for implementation)")
        print("=" * 80)
        
        for sheet_name, sheet_fields in fields.items():
            print(f"\n## {sheet_name}")
            print("```typescript")
            print("interface " + sheet_name.replace(" ", "") + "Data {")
            for field in sheet_fields[:20]:  # Show first 20 fields per sheet
                # Create field name from label
                field_name = field['label'][:50].lower()
                field_name = field_name.replace(' ', '_')
                field_name = field_name.replace('-', '_')
                field_name = field_name.replace('[', '').replace(']', '')
                field_name = field_name.replace('(', '').replace(')', '')
                field_name = ''.join(c if c.isalnum() or c == '_' else '' for c in field_name)
                field_name = field_name[:30]  # Limit length
                
                # Determine TypeScript type
                ts_type = 'string'
                if field['data_type'] == 'number':
                    ts_type = 'number'
                elif field['data_type'] == 'boolean':
                    ts_type = 'boolean'
                elif field['options']:
                    ts_type = f'"{'\" | \"'.join(field['options'][:5])}"'
                
                optional = '?' if not field['required'] else ''
                print(f"  {field_name}{optional}: {ts_type}  // {field['label'][:50]}")
            print("}")
            print("```")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()


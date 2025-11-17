#!/usr/bin/env python3
"""
Extract exact field structure from VSME Excel for Basic Report
"""
import sys
import json
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

def extract_field_structure(ws, sheet_name):
    """Extract field structure with exact cell positions"""
    fields = []
    
    # Basic Report sheets typically have:
    # Column A: Paragraph Reference
    # Column B: Paragraph Guidance Reference  
    # Column C: Field Label/Question
    # Column D onwards: Data fields
    
    for row_idx in range(1, min(ws.max_row + 1, 300)):
        # Check column C for field labels (disclosures)
        label_cell = ws.cell(row=row_idx, column=3)
        
        if not label_cell.value:
            continue
        
        label = str(label_cell.value).strip()
        
        # Skip if too short or not a field
        if len(label) < 10:
            continue
        
        # Look for Basic Module disclosures (B1, B3, B8, B9, B10, B11)
        is_basic = False
        disclosure_code = None
        
        if 'B1' in label or 'B3' in label or 'B8' in label or 'B9' in label or 'B10' in label or 'B11' in label:
            is_basic = True
            # Extract disclosure code
            for code in ['B1', 'B3', 'B8', 'B9', 'B10', 'B11']:
                if code in label:
                    disclosure_code = code
                    break
        
        # Also check for required fields
        is_required = '[Always to be reported]' in label or '[Alw' in label
        is_optional = '[If applicable]' in label or '[If appl' in label
        
        if not is_basic and not is_required and not is_optional:
            # Skip if not clearly a Basic Report field
            continue
        
        # Get paragraph reference from column A
        para_ref = ws.cell(row=row_idx, column=1).value
        para_ref = str(para_ref).strip() if para_ref else None
        
        # Get guidance reference from column B
        guidance_ref = ws.cell(row=row_idx, column=2).value
        guidance_ref = str(guidance_ref).strip() if guidance_ref else None
        
        # Check what type of field this is by looking at data cells
        field_type = 'text'
        data_cells = []
        
        # Check columns D-G for data
        for col_idx in range(4, min(10, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).strip() not in ['-', 'N/A', 'TBD', '']:
                data_cells.append({
                    'col': col_idx,
                    'value': str(cell.value).strip()[:100]
                })
        
        # Determine field type
        if 'year' in label.lower() or 'date' in label.lower():
            field_type = 'date'
        elif 'number' in label.lower() or 'amount' in label.lower() or 'rate' in label.lower():
            field_type = 'number'
        elif any(x in label.lower() for x in ['yes', 'no', 'has', 'does', 'is']):
            field_type = 'boolean'
        elif 'url' in label.lower() or 'link' in label.lower():
            field_type = 'url'
        elif 'email' in label.lower():
            field_type = 'email'
        
        # Look for dropdown options in nearby rows
        options = []
        # Check if there are options listed below (usually in column D)
        for check_row in range(row_idx + 1, min(row_idx + 15, ws.max_row + 1)):
            option_cell = ws.cell(row=check_row, column=4)
            if option_cell.value:
                opt_val = str(option_cell.value).strip()
                # Valid option if it's short and not a number/date pattern
                if 2 <= len(opt_val) <= 50 and opt_val not in ['-', 'N/A']:
                    # Check if it's likely an option vs. a data value
                    if not opt_val.replace('.', '').replace('-', '').isdigit():
                        if opt_val not in options:
                            options.append(opt_val)
                            if len(options) >= 10:  # Limit options
                                break
        
        field = {
            'sheet': sheet_name,
            'row': row_idx,
            'disclosure_code': disclosure_code,
            'paragraph_reference': para_ref,
            'guidance_reference': guidance_ref,
            'label': label,
            'field_type': field_type,
            'required': is_required,
            'optional': is_optional,
            'data_cells': data_cells,
            'options': options if options else None
        }
        
        fields.append(field)
    
    return fields

def main():
    excel_file = "VSME-Digital-Template-1.1.0.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    basic_sheets = [
        'General Information',
        'Environmental Disclosures', 
        'Social Disclosures',
        'Governance Disclosures'
    ]
    
    all_fields = {}
    
    for sheet_name in basic_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            fields = extract_field_structure(ws, sheet_name)
            all_fields[sheet_name] = fields
            
            print(f"\n{'='*80}")
            print(f"{sheet_name}: {len(fields)} fields")
            print(f"{'='*80}")
            
            # Group by disclosure code
            by_code = {}
            for field in fields:
                code = field['disclosure_code'] or 'Other'
                if code not in by_code:
                    by_code[code] = []
                by_code[code].append(field)
            
            for code, code_fields in sorted(by_code.items()):
                print(f"\n{code}: {len(code_fields)} fields")
                for field in code_fields[:5]:  # Show first 5
                    req = "[REQUIRED]" if field['required'] else "[OPTIONAL]" if field['optional'] else ""
                    print(f"  {req} Row {field['row']:3d}: {field['label'][:70]}")
    
    # Save to JSON for reference
    with open('vsme_fields_structure.json', 'w', encoding='utf-8') as f:
        json.dump(all_fields, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'='*80}")
    print("Field structure saved to vsme_fields_structure.json")
    print(f"{'='*80}")
    
    # Generate summary
    total = sum(len(fields) for fields in all_fields.values())
    required = sum(sum(1 for f in fields if f['required']) for fields in all_fields.values())
    
    print(f"\nSummary:")
    print(f"  Total fields: {total}")
    print(f"  Required fields: {required}")
    print(f"  Optional fields: {total - required}")
    
    wb.close()

if __name__ == "__main__":
    main()


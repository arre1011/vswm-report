#!/usr/bin/env python3
"""
Comprehensive extraction of VSME Excel Template structure
Extracts all modules, disclosures, datapoints, and Named Ranges
"""
import sys
import json
import re
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


def extract_all_named_ranges(wb):
    """Extract all Named Ranges with their definitions"""
    named_ranges = {}
    
    for name_str, definition in wb.defined_names.items():
        try:
            # Get the cell reference
            destinations = list(definition.destinations)
            if destinations:
                sheet_name, cell_ref = destinations[0]
                named_ranges[name_str] = {
                    'name': name_str,
                    'reference': f"'{sheet_name}'!{cell_ref}",
                    'sheet': sheet_name,
                    'cellRef': cell_ref
                }
        except Exception as e:
            print(f"Warning: Could not parse named range {name_str}: {e}")
    
    return named_ranges


def extract_table_of_contents(wb):
    """Extract the complete module structure from Table of Contents sheet"""
    ws = wb['Table of Contents & Validation']
    
    structure = {
        'basicModules': [],
        'comprehensiveModules': [],
        'additionalSections': []
    }
    
    current_section = None
    current_module = None
    current_disclosure = None
    
    for row_idx in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        
        if not col_a and not col_b:
            continue
        
        text = str(col_a).strip() if col_a else str(col_b).strip() if col_b else ""
        
        if not text or text == '-':
            continue
        
        # Identify section headers
        if 'Basic Module' in text:
            current_section = 'basic'
            continue
        elif 'Comprehensive Module' in text:
            current_section = 'comprehensive'
            continue
        elif text in ['General Information', 'Environmental Disclosures', 'Social Disclosures', 'Governance Disclosures']:
            # This is a sheet grouping header
            continue
        
        # Check if it's a module (starts with · and has module code)
        if text.startswith('·'):
            module_text = text.replace('·', '').strip()
            
            # Extract module code (B1, B2, C1, etc.)
            module_match = re.match(r'^([BC]\d+)\s*[-–]\s*(.+)$', module_text)
            
            if module_match:
                module_code = module_match.group(1)
                module_name = module_match.group(2).strip()
                
                current_module = {
                    'moduleCode': module_code,
                    'moduleName': module_name,
                    'disclosures': []
                }
                
                if current_section == 'basic':
                    structure['basicModules'].append(current_module)
                elif current_section == 'comprehensive':
                    structure['comprehensiveModules'].append(current_module)
        
        # Check if it's a disclosure (sub-item under a module)
        elif current_module and not text.startswith('·'):
            # This is a disclosure or sub-disclosure
            if text and len(text) > 5:
                current_module['disclosures'].append({
                    'disclosureName': text,
                    'required': 'MISSING VALUE' in str(col_b) or '[Always to be reported]' in text,
                    'optional': '[If applicable]' in text
                })
    
    return structure


def extract_detailed_sheet_structure(ws, sheet_name, named_ranges):
    """Extract detailed field structure from a specific sheet"""
    fields = []
    
    # Track the current module/disclosure context
    current_module = None
    current_disclosure = None
    
    for row_idx in range(1, min(ws.max_row + 1, 300)):
        # Column A: Paragraph Reference
        # Column B: Guidance Reference
        # Column C: Field Label/Description
        # Column D+: Data fields
        
        para_ref = ws.cell(row=row_idx, column=1).value
        guidance_ref = ws.cell(row=row_idx, column=2).value
        label_cell = ws.cell(row=row_idx, column=3)
        
        if not label_cell.value:
            continue
        
        label = str(label_cell.value).strip()
        
        # Skip very short labels
        if len(label) < 10:
            continue
        
        # Extract module code from label
        module_match = re.search(r'\b([BC]\d+)\b', label)
        if module_match:
            current_module = module_match.group(1)
        
        # Determine if required or optional
        is_required = '[Always to be reported]' in label or '[Alw' in label
        is_optional = '[If applicable]' in label
        
        # Try to find matching named range for this field
        matching_ranges = []
        label_keywords = re.findall(r'\b\w{4,}\b', label.lower())
        
        for range_name, range_info in named_ranges.items():
            if range_info['sheet'] == sheet_name:
                # Check if this named range might correspond to this field
                range_name_lower = range_name.lower()
                if any(keyword in range_name_lower for keyword in label_keywords[:5]):
                    matching_ranges.append(range_name)
        
        # Determine field type
        field_type = 'text'
        if 'date' in label.lower() or 'year' in label.lower() or 'month' in label.lower() or 'day' in label.lower():
            field_type = 'date'
        elif 'number' in label.lower() or 'amount' in label.lower() or 'rate' in label.lower() or 'percentage' in label.lower():
            field_type = 'number'
        elif 'yes' in label.lower() or 'no' in label.lower() or 'true' in label.lower() or 'false' in label.lower():
            field_type = 'boolean'
        elif 'select' in label.lower() or 'choose' in label.lower():
            field_type = 'select'
        elif 'table' in label.lower() or 'list' in label.lower():
            field_type = 'table'
        
        field = {
            'sheet': sheet_name,
            'row': row_idx,
            'module': current_module,
            'label': label,
            'paragraphReference': str(para_ref).strip() if para_ref else None,
            'guidanceReference': str(guidance_ref).strip() if guidance_ref else None,
            'fieldType': field_type,
            'required': is_required,
            'optional': is_optional,
            'potentialNamedRanges': matching_ranges[:5] if matching_ranges else []
        }
        
        fields.append(field)
    
    return fields


def main():
    excel_file = "VSME-Digital-Template-1.1.0.xlsx"
    
    print("=" * 80)
    print("VSME Complete Structure Extraction")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Step 1: Extract all Named Ranges
        print("\n1. Extracting Named Ranges...")
        named_ranges = extract_all_named_ranges(wb)
        print(f"   Found {len(named_ranges)} named ranges")
        
        # Step 2: Extract Table of Contents structure
        print("\n2. Extracting Table of Contents structure...")
        toc_structure = extract_table_of_contents(wb)
        print(f"   Found {len(toc_structure['basicModules'])} Basic Modules")
        print(f"   Found {len(toc_structure['comprehensiveModules'])} Comprehensive Modules")
        
        # Step 3: Extract detailed structure from each sheet
        print("\n3. Extracting detailed field structure from sheets...")
        
        sheets_to_analyze = [
            'General Information',
            'Environmental Disclosures',
            'Social Disclosures',
            'Governance Disclosures'
        ]
        
        detailed_structure = {}
        for sheet_name in sheets_to_analyze:
            if sheet_name in wb.sheetnames:
                print(f"   Analyzing {sheet_name}...")
                ws = wb[sheet_name]
                fields = extract_detailed_sheet_structure(ws, sheet_name, named_ranges)
                detailed_structure[sheet_name] = fields
                print(f"     Found {len(fields)} fields")
        
        # Step 4: Compile complete structure
        print("\n4. Compiling complete structure...")
        complete_structure = {
            'metadata': {
                'sourceFile': excel_file,
                'totalNamedRanges': len(named_ranges),
                'sheets': list(detailed_structure.keys())
            },
            'namedRanges': named_ranges,
            'tableOfContents': toc_structure,
            'detailedFields': detailed_structure
        }
        
        # Step 5: Save to JSON
        output_file = 'vsme-complete-structure.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(complete_structure, f, indent=2, ensure_ascii=False)
        
        print(f"\n✓ Complete structure saved to {output_file}")
        
        # Summary statistics
        print("\n" + "=" * 80)
        print("SUMMARY")
        print("=" * 80)
        print(f"Total Named Ranges: {len(named_ranges)}")
        print(f"\nBasic Modules ({len(toc_structure['basicModules'])}):")
        for module in toc_structure['basicModules']:
            print(f"  {module['moduleCode']}: {module['moduleName']}")
            print(f"    └─ {len(module['disclosures'])} disclosures")
        
        print(f"\nComprehensive Modules ({len(toc_structure['comprehensiveModules'])}):")
        for module in toc_structure['comprehensiveModules']:
            print(f"  {module['moduleCode']}: {module['moduleName']}")
            print(f"    └─ {len(module['disclosures'])} disclosures")
        
        print(f"\nDetailed Fields by Sheet:")
        total_fields = 0
        for sheet_name, fields in detailed_structure.items():
            total_fields += len(fields)
            required = sum(1 for f in fields if f['required'])
            optional = sum(1 for f in fields if f['optional'])
            print(f"  {sheet_name}: {len(fields)} fields ({required} required, {optional} optional)")
        
        print(f"\nTotal fields extracted: {total_fields}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

